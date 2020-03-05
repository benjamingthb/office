import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

"""write excel"""

#创建一个工作表
wb = Workbook() 
dest_filename = 'empty_book.xlsx'

#激活sheet
ws1 = wb.active 
ws1.title = "range names"
#向sheet中写入0-599，写39行
for row in range(1,40): 
    ws1.append(range(600))
#合并单元格A2倒D2，值取A2
ws1.merge_cells('A2:D2')
#取消合并，主意这时B2:D2的值为空
ws1.unmerge_cells('A2:D2')

ws2 = wb.create_sheet(title="Pi")
#向F5写入值
ws2['F5'] = 3.14
#写入当前日期
ws2['A1'] = datetime.datetime.now()
#写入公式
ws2['B1'] = "=SUM(1,1)"

ws3 = wb.create_sheet(title="Data")
#向10-19行，27-53行写入值，值为列的名字
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

wb.save(filename = dest_filename)

"""read excel"""
wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)
